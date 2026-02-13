from __future__ import annotations
import csv
from pathlib import Path
from typing import Dict, List, Optional

from .system_health import HealthSnapshot
from .log_parser import LogParseResult
from .utils import ensure_dir, safe_str

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None  # type: ignore

def _autosize_columns(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in list(col)[:200]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)

def export_csv(
    out_dir: Path,
    summary: Dict[str, str],
    snapshot: Optional[HealthSnapshot],
    log_result: Optional[LogParseResult],
) -> List[Path]:
    ensure_dir(out_dir)
    created: List[Path] = []

    p_sum = out_dir / "summary.csv"
    with p_sum.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["key", "value"])
        for k, v in summary.items():
            w.writerow([k, v])
    created.append(p_sum)

    if snapshot is not None:
        p_sys = out_dir / "system_health.csv"
        with p_sys.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([
                "timestamp", "hostname", "os", "os_version", "mount",
                "disk_total_gb", "disk_used_gb", "disk_free_gb", "disk_used_percent",
                "mem_total_gb", "mem_used_gb", "mem_free_gb", "mem_used_percent",
                "cpu_cores_logical", "cpu_load_percent"
            ])
            for d in snapshot.disks:
                w.writerow([
                    snapshot.timestamp, snapshot.hostname, snapshot.os, snapshot.os_version,
                    d.mount, safe_str(d.total_gb), safe_str(d.used_gb), safe_str(d.free_gb), safe_str(d.used_percent),
                    safe_str(snapshot.memory.total_gb), safe_str(snapshot.memory.used_gb), safe_str(snapshot.memory.free_gb), safe_str(snapshot.memory.used_percent),
                    safe_str(snapshot.cpu.cores_logical), safe_str(snapshot.cpu.load_percent),
                ])
        created.append(p_sys)

        p_svc = out_dir / "services.csv"
        with p_svc.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["name", "status", "detail"])
            for s in snapshot.services:
                w.writerow([s.name, s.status, s.detail])
        created.append(p_svc)

    if log_result is not None:
        p_log = out_dir / "log_findings.csv"
        with p_log.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["file", "total_lines", "matched_lines"])
            w.writerow([log_result.file, log_result.total_lines, log_result.matched_lines])
            w.writerow([])
            w.writerow(["keyword", "count"])
            for kw, c in sorted(log_result.by_keyword.items(), key=lambda x: (-x[1], x[0])):
                w.writerow([kw, c])
            w.writerow([])
            w.writerow(["sample_line_no", "keyword", "line"])
            for s in log_result.samples:
                w.writerow([s.line_no, s.keyword, s.line])
        created.append(p_log)

    return created

def export_xlsx(
    out_file: Path,
    summary: Dict[str, str],
    snapshot: Optional[HealthSnapshot],
    log_result: Optional[LogParseResult],
) -> Path:
    if Workbook is None:
        raise RuntimeError("openpyxl is not available; install dependencies or use CSV export.")

    out_file.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["key", "value"])
    for k, v in summary.items():
        ws.append([k, v])
    _autosize_columns(ws)

    if snapshot is not None:
        ws2 = wb.create_sheet("SystemHealth")
        ws2.append([
            "timestamp", "hostname", "os", "os_version", "mount",
            "disk_total_gb", "disk_used_gb", "disk_free_gb", "disk_used_percent",
            "mem_total_gb", "mem_used_gb", "mem_free_gb", "mem_used_percent",
            "cpu_cores_logical", "cpu_load_percent"
        ])
        for d in snapshot.disks:
            ws2.append([
                snapshot.timestamp, snapshot.hostname, snapshot.os, snapshot.os_version,
                d.mount, d.total_gb, d.used_gb, d.free_gb, d.used_percent,
                snapshot.memory.total_gb, snapshot.memory.used_gb, snapshot.memory.free_gb, snapshot.memory.used_percent,
                snapshot.cpu.cores_logical, snapshot.cpu.load_percent
            ])
        _autosize_columns(ws2)

        ws3 = wb.create_sheet("Services")
        ws3.append(["name", "status", "detail"])
        for s in snapshot.services:
            ws3.append([s.name, s.status, s.detail])
        _autosize_columns(ws3)

    if log_result is not None:
        ws4 = wb.create_sheet("LogFindings")
        ws4.append(["file", log_result.file])
        ws4.append(["total_lines", log_result.total_lines])
        ws4.append(["matched_lines", log_result.matched_lines])
        ws4.append([])
        ws4.append(["keyword", "count"])
        for kw, c in sorted(log_result.by_keyword.items(), key=lambda x: (-x[1], x[0])):
            ws4.append([kw, c])
        ws4.append([])
        ws4.append(["sample_line_no", "keyword", "line"])
        for s in log_result.samples:
            ws4.append([s.line_no, s.keyword, s.line])
        _autosize_columns(ws4)

    wb.save(out_file)
    return out_file
